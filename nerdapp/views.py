import datetime
from django.urls import reverse
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.views import View
from .models import *
from .forms import *
from django.contrib.auth import authenticate, login
from django.contrib import messages
from paypal.standard.forms import PayPalPaymentsForm
from django.conf import settings
import uuid
from django.db.models import Q
from django.core.mail import send_mail
from random import randrange
from django.db.models import Sum
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from paypal.standard.ipn.signals import valid_ipn_received
from django.views.decorators.csrf import csrf_exempt
from paypal.standard.ipn.views import ipn
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from nerdcoudstore.settings import EMAIL_HOST_USER



def index(request):
    productos = Producto.objects.all().order_by("-fecha_creacion")[:10]

    data = {
        'productos': productos
    }

    return render(request, 'nerdapp/index.html', data)

def page2(request):
    return render(request, 'nerdapp/page2.html')

def testPaypal(request):
    return render(request, 'payment/testpaypal.html')

def lista_categorias(request):
    categorias = Categoria.objects.all()
    #print(categorias)
    return render(request, 'nerdapp/lista_categorias.html', {'categorias': categorias})

def listForo(request):
    return render(request, 'foro/listForo.html')

def signup(request):
    data = {
        'form': CustomUserCreationForm
    }
    if request.method == 'POST':
        formulario = CustomUserCreationForm(data=request.POST)
        if formulario.is_valid():
            aux_user = formulario.save()
            usuario = Usuario.objects.create(
                id_usuario = aux_user.id,
                nombre_usuario = aux_user.username,
                nombre = aux_user.first_name,
                apellido = aux_user.last_name,
                correo = aux_user.email,        
            )
            carrito = Carrito.objects.create(
                usuario_id_usuario = usuario,
                estado_pago = 'pendiente',
                total_carrito = 0
            )
            carrito.save()
            subject = "Usuario/a Creado" 
            message = "Estimado/a usuario/a" + " " + aux_user.username + ", su cuenta de NerdCoudStore ha sido creada."
            email = aux_user.email
            recipient_list = [email]
            send_mail(subject, message, EMAIL_HOST_USER, recipient_list, fail_silently=True) 
            usuario.save()

            user = authenticate(username=formulario.cleaned_data['username'], password=formulario.cleaned_data['password1'])
            login(request, user)
            messages.success(request,"cuenta creada correctamente")
            return redirect(to="index")
        data["form"] = formulario

        return render(request, 'registration/signup.html', data)

    if request.method == 'POST':
        formulario = CustomUserCreationForm(data=request.POST)
        if formulario.is_valid():
            aux_user = formulario.save()
            usuario = Usuario.objects.create(
                id_usuario = aux_user.id,
                nombre_usuario = aux_user.username,
                nombre = aux_user.first_name,
                apellido = aux_user.last_name,
                correo = aux_user.email,        
            )
            carrito = Carrito.objects.create(
                usuario_id_usuario = usuario,
                estado_pago = 'pendiente',
                total_carrito = 0
            )
            carrito.save()
            subject = "Usuario/a Creado" 
            message = "Estimado/a usuario/a" + " " + aux_user.username + ", su cuenta de NerdCoudStore ha sido creada."
            email = aux_user.email
            recipient_list = [email]
            send_mail(subject, message, EMAIL_HOST_USER, recipient_list, fail_silently=True) 
            usuario.save()

            user = authenticate(username=formulario.cleaned_data['username'], password=formulario.cleaned_data['password1'])
            login(request, user)
            messages.success(request,"cuenta creada correctamente")
            return redirect(to="index")
        data["form"] = formulario

    return render(request, 'registration/signup.html', data)

def agregarSubasta(request):
    data = {
        'form': SubastaForm()
    }
    if request.method == 'POST':
        formulario = SubastaForm(data=request.POST, files=request.FILES)
        #print("formulario ",formulario)
        if formulario.is_valid:
            subasta = formulario.save()  # Guarda la subasta y obtén la instancia
            usuario = Usuario.objects.get(id_usuario=request.user.id)  # Obtiene el usuario actual
            usuario_subasta = Usuario_subasta.objects.create(
                usuario_id_usuario=usuario,  # Obtiene el id del usuario actual
                subasta_id_subasta=subasta # Usa el id de la subasta recién creada
            )

            subject = "Subasta Creada" 
            message = "Estimado/a " + usuario.nombre_usuario + ", su subasta '" + subasta.nombre + "' ha sido creada con un valor inicial de " + "$" + str(subasta.precio_inicial) + " CLP."
            email = usuario.correo
            recipient_list = [email]
            html_message = f"""<p>{message}</p><img src="https://i.imgur.com/wSs6Cnr.png" title="source: imgur.com" />
            <p>Encuéntranos en Avenida Concha Y Toro, Av. San Carlos 1340</p>"""
            send_mail(subject, message, EMAIL_HOST_USER, recipient_list, fail_silently=True, html_message=html_message)

            usuario_subasta.save()  # Guarda la relación en UsuarioSubasta
            data['mensaje']="guardado correctamente"
            return redirect('listSubastas')
        else:
            data["form"] = formulario
    return render(request, 'subasta/agregarSubasta.html',data)

class ListarYParticiparSubastas(View):
    def get(self, request, *args, **kwargs):
        subastas = Subasta.objects.all().order_by('-fecha_inicio')
        form = ParticiparSubastaForm()
        context = {
            'subastas': subastas,
            'form': form,
        }
        return render(request, 'subasta/listSubastas.html', context)
    def post(self, request, *args, **kwargs):
        formulario = ParticiparSubastaForm(data=request.POST, files=request.FILES)
        subastas = Subasta.objects.all().order_by('-fecha_inicio')
        context = {
            'subastas': subastas,
            'form': formulario,
        }
        usuario =request.POST.get('usuario_id')           
        subasta = request.POST.get('subasta_id')
        montoS = request.POST.get('monto')
        subastaClass = Subasta.objects.get(id_subasta=subasta)
        """print("Usuario:", usuario)
        print("Subasta:", subasta)
        print("precio_inicial:", subastaClass.precio_inicial)
        print("precio_mas_alto:", subastaClass.precio_mas_alto)
        print("Monto:", montoS)"""
        if formulario.is_valid and int(montoS) > int(subastaClass.precio_inicial) and int(montoS) >  int(subastaClass.precio_mas_alto) :
            usuarioClass = Usuario.objects.get(id_usuario=request.user.id)      
            participarSubastaUsuario = ParticiparSubasta.objects.create(
                usuario_id_usuario_id=usuarioClass,
                subasta_id_subasta_id=subastaClass,
                monto=montoS
            )
            participarSubastaUsuario.save()
            subastaClass.precio_mas_alto = montoS

            subject = "Participación en Subasta Exitosa"
            message = f"Estimado/a {usuarioClass.nombre_usuario}, su participación en la subasta '{subastaClass.nombre}' ha sido exitosa."
            email = usuarioClass.correo
            recipient_list = [email]

            html_message = f"""
            <p>{message}</p>
            <p>Ha ofertado con éxito en la subasta. El monto de su oferta es de ${montoS} CLP.</p>
            <img src="https://i.imgur.com/wSs6Cnr.png" alt="Firma">
            <p>Encuéntranos en Avenida Concha Y Toro, Av. San Carlos 1340</p>
            """

            send_mail(subject, message, settings.EMAIL_HOST_USER, recipient_list, fail_silently=True, html_message=html_message)
            
            subastaClass.save()
            return redirect(reverse('participacionSubasta', args=[subasta, montoS]))
        elif  int(montoS) < int(subastaClass.precio_inicial) :
            context['mensaje'] = 'El monto ingresado es menor a la oferta mas alta'
        else:
            context['mensaje'] = 'Hubo un error al procesar el formulario. Por favor, inténtelo de nuevo.'

        return render(request, 'subasta/listSubastas.html', context)

def participacionSubasta(request, subasta_id, monto):
    #print("Subasta:", subasta_id)
    #print("Monto:", monto)
    try:
        subasta = Subasta.objects.get(id_subasta=subasta_id)
        context = {
            'subasta': subasta,
            'descripcion' : subasta.descripcion,
            'fecha_termino' : subasta.fecha_termino,
            'hora_termino' : subasta.hora_termino,
            'monto': monto,
        }
        return render(request, 'subasta/participacionSubasta.html', context)
    except Subasta.DoesNotExist:
        return HttpResponse("Subasta no encontrada.")

def ProductView(request):

    get_productos = Producto.objects.all()

    return render(request, 'product.html', {'productos': get_productos})

def CheckOut(request, id_producto):

    productos = Producto.objects.get(id_producto=id_producto)

    clp_a_usd = 0.0011
    
    precio_clp = float(productos.precio)

    monto_usd = precio_clp * clp_a_usd

    host = request.get_host()

    paypal_checkout = {
        'business': settings.PAYPAL_RECEIVER_EMAIL,
        'amount': monto_usd,
        'item_name': productos.nombre,
        'invoice': uuid.uuid4(),
        'currency_code': 'USD',
        'notify_url': f"http://{host}{reverse('paypal-ipn')}",
        'return_url': f"http://{host}{reverse('payment-success', kwargs = {'id_producto': productos.id_producto})}",
        'cancel_url': f"http://{host}{reverse('payment-failed', kwargs = {'id_producto': productos.id_producto})}",
    }

    paypal_payment = PayPalPaymentsForm(initial=paypal_checkout)

    context = {
        'productos': productos,
        'paypal': paypal_payment
    }

    return render(request, 'nerdapp/checkout.html', context)

def PaymentSuccessful(request, id_carrito):
    carrito = Carrito.objects.get(id_carrito=id_carrito)
    usuario = Usuario.objects.get(id_usuario=request.user.id)

    subject = "Compra Exitosa"
    message = f"Estimado/a  ¡su compra de {carrito.total_carrito} ha sido exitosa!"
    email = usuario.correo
    recipient_list = [email]

    html_message = f"""<p>{message}</p><img src="https://i.imgur.com/wSs6Cnr.png" alt="Firma">
    <p>Encuéntranos en Avenida Concha Y Toro, Av. San Carlos 1340</p>"""

    send_mail(subject, message, settings.EMAIL_HOST_USER, recipient_list, fail_silently=True, html_message=html_message)
    
    carrito.estado_pago='pagado'
    carrito.save()
    carritoNuevo = Carrito.objects.create(
                usuario_id_usuario = usuario,
                estado_pago = 'pendiente',
                total_carrito = 0
            )
    carritoNuevo.save()
    return render(request, 'nerdapp/payment-success.html')

def paymentFailed(request, id_carrito):

    producto = Carrito.objects.get(id_carrito=id_carrito)
    usuario = request.user

    subject = "Compra fallida"
    message = f"Estimado/a {usuario.nombre_usuario}, su compra de {producto.total_carrito} no se ha completado"
    email = usuario.correo
    recipient_list = [email]

    html_message = f"""<p>{message}</p><img src="https://i.imgur.com/wSs6Cnr.png" alt="Firma">
    <p>Encuéntranos en Avenida Concha Y Toro, Av. San Carlos 1340</p>"""

    send_mail(subject, message, settings.EMAIL_HOST_USER, recipient_list, fail_silently=True, html_message=html_message)

    return render(request, 'nerdapp/payment-failed.html', {'carrito': producto})

class listarYComentarForo(View):
    def get(self, request, *args, **kwargs):
        publicaciones = Publicacion.objects.filter(estado_publicacion='activo').order_by('-fecha_publicacion')
        comentarios = Comentario.objects.filter(estado_comentario='activo').order_by('-fecha_comentario')
        context = {
                'publicaciones': publicaciones,
                'comentarios': comentarios
            }
        return render(request, 'foro/listForo.html', context)
    def post(self, request, *args, **kwargs):
        formulario = ComentarForo(data=request.POST, files=request.FILES)
        foro= request.POST.get('publicacion_id_publicacion')
        #print(foro)
        if formulario.is_valid():
            formulario.save()
            return redirect(reverse('participacionForo', args=[foro]))
        
def participacionForo(request, id_publicacion):
    #print("Foro:", id_publicacion)
    try:
        foro = Publicacion.objects.get(id_publicacion=id_publicacion)
        context = {
            'foro': foro.titulo_publicacion,
        }
        return render(request, 'foro/participacionForo.html', context)
    except Publicacion.DoesNotExist:
        return HttpResponse("Foro no encontrado.")
   
def agregarForo(request):
    data = {'form': ForoForm()}
    if request.method =="POST":
        formulario = ForoForm(data=request.POST, files=request.FILES)
        #print(formulario)
        if formulario.is_valid():
            #print("formulario valido")

            # Obtener el usuario actual
            usuario = request.user

            # Envía un correo de confirmación
            subject = "Foro Creado Exitosamente"
            message = f"Estimado/a {usuario.username}, su foro ha sido creado exitosamente." 
            email = usuario.email
            recipient_list = [email]

            html_message = f"""
            <p>{message}</p>
            <img src="https://i.imgur.com/wSs6Cnr.png" alt="Firma" style="width: 100%">
            <p>Encuéntranos en Avenida Concha Y Toro, Av. San Carlos 1340</p>
            """

            send_mail(subject, message, settings.EMAIL_HOST_USER, recipient_list, fail_silently=True, html_message=html_message)

            data['mensaje'] = "Guardado correctamente"

            formulario.save()
        #usuario = Usuario.objects.get(id_usuario=request.user.id)  # Obtiene el usuario actual
            data['mensaje']="guardado correctamente"
        else:
        #print("formulario no valido")
            data["form"] = formulario        
    return render(request, 'foro/agregarForo.html',data)

def reportarComentario(request, comentarioID, publicacionID):
    if(comentarioID != 0):
        print(comentarioID, publicacionID)
        comentario = Comentario.objects.get(id_comentario=comentarioID)
        comentario.estado_comentario = 'denunciado'
        comentario.save()
        
        print(comentario)
        # Renderiza la plantilla directamente, no es necesario pasar args
        
    else:
        print(comentarioID, publicacionID)
        publicacion = Publicacion.objects.get(id_publicacion=int(publicacionID))
        publicacion.estado_publicacion = 'denunciado'
        publicacion.save()
        # Renderiza la plantilla directamente, no es necesario pasar args
    context={
            'publicacionID':publicacionID
        }
    return render(request, 'foro/reportarForo.html', context)
    
def vistaVenta(request):
    diccAlias, diccNombre = diccProductos()
    diccIdNombreCant = [(str(id), nombre, cantidad, precio, total) for id, nombre, cantidad, precio, total in diccNombre]
    totalVentas = 0
    for t in diccIdNombreCant:
        totalVentas=totalVentas+t[4]
    context = {
        'diccIdNombreCant': diccIdNombreCant,
        'totalVentaMes':totalVentas
    }

    return render(request, 'informe/vistaVenta.html', context)

def get_chart(request):
    diccAlias,diccNombre = diccProductos()
    diccionario_ordenado=diccAlias
    now = datetime.now()
    current_month = now.month
    current_year = now.year
    nombres_productos = list(diccionario_ordenado.keys())
    cantidad_prodcutos = list(diccionario_ordenado.values())
    #print(cantidad_prodcutos)
    #print(nombres_productos)
    series = cantidad_prodcutos

    chart={
        'xAxis':[
            {
                'type' : "category",
                'data' : nombres_productos,
                'name': 'Productos',  # Título del eje X
                'axisLabel': {
                    'rotate': 45,  # Rotar etiquetas del eje X para mejor legibilidad
                },
            }
        ],
        'yAxis':[
            {
                'type' : "value",
                'name': 'Cantidad Vendida en '+str(current_month)+'/'+str(current_year),
            }
        ],
        'series':[
            {
            'data':series,
            'type': 'bar',
            'showBackground': 'true',
            'backgroundStyle': {
                'color': '#6264eb21'
            }
            }
        ]
    }
    return JsonResponse(chart)

def diccProductos():
    productos = Producto.objects.all().values()
    carritoUsuarios = Carrito.objects.filter(
            estado_pago="pagado"
        ).values()
    cont=0
    lisCarritoProductos =[]
    while(cont<len(carritoUsuarios)):
        carritosProductos = CarritoProducto.objects.filter(
            id_carrito_id=carritoUsuarios[cont]['id_carrito']
        ).values()
        lisCarritoProductos.extend(carritosProductos)
        cont+=1
    now = datetime.now()
    current_month = now.month
    current_year = now.year
    ventas_por_mes = Venta.objects.filter(
        fecha_venta__month =current_month,
        fecha_venta__year =current_year
    ).values()
    diccProductos ={str(producto['id_producto'])+"-"+producto['nombre'][:5]: 0 for producto in productos}
    diccNombres = {str(producto['id_producto'])+"|"+producto['nombre']: 0 for producto in productos}
    #diccNombres=[]
    for venta in ventas_por_mes:
        for carpro in lisCarritoProductos:
            if venta['id_carrito_id_id'] == carpro['id_carrito_id_id']:
                producto = Producto.objects.filter(
                    id_producto=carpro['id_producto_id_id']
                ).values()[0]
                diccProductos.update({str(producto['id_producto'])+"-"+producto['nombre'][:5]:diccProductos[str(producto['id_producto'])+"-"+producto['nombre'][:5]]+int(carpro['cantidad_producto'])})
                #diccNombres.append(producto['id_producto'],producto['nombre'],)
                diccNombres.update({str(producto['id_producto'])+"|"+producto['nombre']:diccNombres[str(producto['id_producto'])+"|"+producto['nombre']]+int(carpro['cantidad_producto'])})
        #diccNombres.append([producto['id_producto'],producto['nombre'],diccProductos[producto['nombre']]])
    #print(diccNombres)
    diccionario_ordenado = dict(sorted(diccProductos.items(), key=lambda item: item[1], reverse=True))
    MatAux=[] 
    for key,value in diccionario_ordenado.items():
        id,alias=key.split("-", 1)
        producto = Producto.objects.get(id_producto=id)
        #print('producto ',producto)
        MatAux.append([id,producto.nombre,value,producto.precio, value*producto.precio])
    #print(MatAux)
    diccionario_Nombres=MatAux
    #diccionario_Nombres = dict(sorted(diccNombres.items(), key=lambda item: item[1], reverse=True))
    #diccionario_Nombres = diccNombres
    #print(diccionario_Nombres)
    #{'3-Cojín': 2, '6-Manta': 1, '2-Tazón': 0, '4-Nicke': 0}
    return (diccionario_ordenado,diccionario_Nombres)

def descargar_pdf(request):
    diccAlias, diccNombre = diccProductos()
    now = datetime.now()
    current_month =str(now.month) 
    current_year = str(now.year)
    fn="informe_ventas_"+current_month+"_"+current_year+".pdf"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fn}"'

    # Crear el objeto PDF usando reportlab
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()

    # Crear la tabla y definir estilos
    data = [['ID', 'Nombre', 'Cantidad', 'Precio unitario', 'Total Vendido']] + diccNombre
    table_style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                              ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                              ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                              ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                              ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                              ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                              ('GRID', (0, 0), (-1, -1), 1, colors.black)])

    table = Table(data, style=table_style)
    doc.build([Paragraph("Informe de Ventas "+current_month+" / "+current_year, styles['Title']), table])

    return response

def descargar_excel(request):
    diccAlias, diccNombre = diccProductos()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="informe_ventas.xlsx"'

    # Crear el objeto Excel usando openpyxl
    wb = Workbook()
    ws = wb.active

    # Agregar encabezados
    ws.append(['ID', 'Nombre', 'Cantidad', 'Precio unitario', 'Total Vendido'])

    # Agregar datos
    for id, nombre, cantidad, precio, total in diccNombre:
        ws.append([id, nombre, cantidad, precio, total])

    # Guardar el libro de trabajo
    wb.save(response)

    return response

def infoBoletas(id_usuario):
    carritosUsuario = Carrito.objects.filter(
        usuario_id_usuario=id_usuario,
        estado_pago="pagado"
    ).values()
    listaVentaProducto=[]
    for c in carritosUsuario:
        ventaCarrito = Venta.objects.get(
            id_carrito_id=c['id_carrito']
        )
        d=str(ventaCarrito.fecha_venta.day)
        m=str(ventaCarrito.fecha_venta.month)
        y=str(ventaCarrito.fecha_venta.year)
        fechaVenta = d+'/'+m+'/'+y
        prodCar = CarritoProducto.objects.filter(
            id_carrito_id=c['id_carrito']
        ).values()

        listProductosCarritoUsuario=[]
        for pc in prodCar:
            producto = Producto.objects.get(id_producto=pc['id_producto_id_id'])
            #print(producto.precio)

            listProductosCarritoUsuario.append([producto.nombre, producto.precio,pc['cantidad_producto'],producto.precio*pc['cantidad_producto'],c['id_carrito'],ventaCarrito.id_venta,fechaVenta, ventaCarrito.total_venta])
        #print(prodCar) #productos de carrito
        listaVentaProducto.append(listProductosCarritoUsuario)
    return listaVentaProducto

def descargarBoleta_pdf(request, boleta_id):
    listaVentaProducto = infoBoletas(request.user.id)
    now = datetime.now()
    current_month = str(now.month)
    current_year = str(now.year)
    fn = "boleta_" + str(boleta_id) + ".pdf"
    #print(listaVentaProducto)
    flattened_data = [item for sublist in listaVentaProducto for item in sublist]
    dataBoleta = [i[:4] for i in flattened_data]
    infoBoleta = [i[4:] for i in flattened_data][0]
    print(infoBoleta)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()

    # Ajustar el margen entre el subtítulo y la tabla
    styles.add(ParagraphStyle(name='TableParagraph', parent=styles['Normal'], spaceAfter=12))

    boleta_title = f"Boleta - {infoBoleta[1]}"
    boleta_subTitle = f"Fecha Venta: {infoBoleta[2]}, Total Venta: {infoBoleta[3]}\n"
    data = [['Nombre', 'Precio Unitario', 'Cantidad', 'Precio Total']] + dataBoleta
    table_style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                              ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                              ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                              ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                              ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                              ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                              ('GRID', (0, 0), (-1, -1), 1, colors.black)])

    table = Table(data, style=table_style)
    doc.build([Paragraph(boleta_title, styles['Title']),
               Paragraph(boleta_subTitle, styles['TableParagraph']),
                 table])

    return response   

def descargarExcelBoletas(request):
    listaVentaProducto = infoBoletas(request.user.id)
    #print(listaVentaProducto)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="informe_ventas.xlsx"'

    # Crear el objeto Excel usando openpyxl
    wb = Workbook()
    ws = wb.active
    ws.append(['ID', 'Nombre', 'Cantidad', 'Precio unitario', 'Total Vendido', 'ID Carrito', 'ID Boleta', 'Fecha Venta', 'Total Boletas'])

    # Agregar datos
        # ...
    for venta in listaVentaProducto:
        count=1
        for nombre, cantidad, precio, total, idC, idB, fec, totalB in venta:
            ws.append([count, nombre, cantidad, precio, total, idC, idB, fec, totalB])
            count+=1
    # ...

    # Guardar el libro de trabajo
    wb.save(response)

    return response

def vistaBoleta(request):
    listaVentaProducto= infoBoletas(request.user.id)
    #print(listaVentaProducto)#lista de ventas con detalle de los productos
    context = {
        "listaVentaProducto":listaVentaProducto
    }
    return render(request, 'informe/vistaBoleta.html', context)

def buscar(request):
    query = request.GET.get('q')

    if query:
        # Realiza la búsqueda en tu modelo (ajusta esto según tu modelo)
        resultados = Producto.objects.filter(nombre__icontains=query)
    else:
        resultados = []

    return render(request, 'busqueda/busqueda.html', {'resultados': resultados, 'query': query})

def buscarSubasta(request):
    query = request.GET.get('q')

    if query:
        # Realiza la búsqueda en tu modelo (ajusta esto según tu modelo)
        resultados = Subasta.objects.filter(nombre__icontains=query)
    else:
        resultados = []

    return render(request, 'busqueda/busquedaSubasta.html', {'subastas': resultados, 'query': query})


def videojuegos(request):

    productos_videojuegos = Producto.objects.filter(categoria_id_categoria='1')


    return render(request, 'nerdapp/videojuegos.html', {'productos': productos_videojuegos})

def mangas(request):

    productos_mangas = Producto.objects.filter(categoria_id_categoria='2')


    return render(request, 'nerdapp/mangas.html', {'productos': productos_mangas})

def animes(request):

    productos_animes = Producto.objects.filter(categoria_id_categoria='3')


    return render(request, 'nerdapp/animes.html', {'productos': productos_animes})


def accesorios(request):

    productos_accesorios = Producto.objects.filter(categoria_id_categoria='4')


    return render(request, 'nerdapp/accesorios.html', {'productos': productos_accesorios})

def agregar_al_carrito(request, id_producto):
    # Obtén el producto con el ID proporcionado o redirige si no existe
    try:
        producto = Producto.objects.get(pk=id_producto)
    except Producto.DoesNotExist:
        return render(request, "Producto no encontrado", status=404)

    # Obtén el carrito del usuario actual o crea uno si no existe
    carrito, creado = Carrito.objects.get_or_create(usuario_id_usuario=request.user.id, estado_pago='pendiente')
    print(carrito)
    # Busca el producto en el carrito
    carrito_producto_existente = CarritoProducto.objects.filter(
        id_carrito_id=carrito,
        id_producto_id=producto
    ).first()

    if carrito_producto_existente:
        # Si el producto ya está en el carrito, incrementa la cantidad
        carrito_producto_existente.cantidad_producto += 1
        carrito_producto_existente.total_por_producto = carrito_producto_existente.cantidad_producto * producto.precio
        carrito_producto_existente.save()
    else:
        # Si el producto no está en el carrito, crea un nuevo CarritoProducto
        nuevo_carrito_producto = CarritoProducto(
            id_producto_id=producto,
            cantidad_producto=1,
            total_por_producto=producto.precio,
            id_carrito_id=carrito
        )
        nuevo_carrito_producto.save()

    # Actualiza el total del carrito
    carrito.total_carrito += producto.precio
    carrito.save()
  
    return render(request, 'carrito/agregado_al_carrito.html')  

def ver_carrito(request):
    carrito = Carrito.objects.get(usuario_id_usuario=request.user.id, estado_pago='pendiente')
    carrito_items = CarritoProducto.objects.filter(id_carrito_id__usuario_id_usuario=request.user.id, id_carrito_id__estado_pago='pendiente')
    carrito_total = carrito_items.aggregate(Sum('total_por_producto'))['total_por_producto__sum'] or 0

    clp_a_usd = 0.0011
    monto_usd = carrito_total * clp_a_usd

    host = request.get_host()

    
    paypal_checkout = {
        'business': settings.PAYPAL_RECEIVER_EMAIL,
        'amount': monto_usd,
        'item_name': 'Compra en Mi Tienda',
        'invoice': uuid.uuid4(),
        'currency_code': 'USD',
        'notify_url': f"http://{host}{reverse('paypal-ipn')}",
        'return_url': f"http://{host}/payment-success/{carrito.id_carrito}", 
    }

    paypal_payment = PayPalPaymentsForm(initial=paypal_checkout)

    context = {
        'carrito_items': carrito_items,
        'carrito_total': carrito_total,
        'paypal': paypal_payment,
    }

    return render(request, 'carrito/ver_carrito.html', context)

def eliminar_del_carrito(request, id_carrito_producto):
    try:
        # Obtén el CarritoProducto con el ID proporcionado
        carrito_producto = CarritoProducto.objects.get(pk=id_carrito_producto)
    except CarritoProducto.DoesNotExist:
        # Si no se encuentra el CarritoProducto, puedes redirigir o manejar el error de alguna manera
        return redirect('nombre_de_tu_vista_de_error')

    # Resta el total del carrito antes de eliminar el producto
    carrito = carrito_producto.id_carrito_id
    carrito.total_carrito -= carrito_producto.total_por_producto
    carrito.save()

    # Elimina el CarritoProducto
    carrito_producto.delete()

    # Redirige a la vista del carrito
    return redirect('ver_carrito')

def paypal_ipn(request):

    usuario=Usuario.objects.get(id_usuario=request.user.id)
    # Procesa la notificación de pago de PayPal y actualiza el estado del carrito
    if request.method == "POST" and request.POST.get("txn_id"):
        ipn(request)
        carritoAnterior = Carrito.objects.get(usuario_id_usuario=usuario.id_usuario)
        carritoAnterior.estado_pago='pagado'
        carritoAnterior.save()
        carritoNuevo = Carrito.objects.create(
                    usuario_id_usuario = usuario,
                    estado_pago = 'pendiente',
                    total_carrito = 0
                )
        carritoNuevo.save()
        # Agrega aquí la lógica para actualizar el estado del carrito después de la notificación de PayPal

    # Devuelve una respuesta adecuada a PayPal
    return HttpResponse("OK")

"""def manejar_notificacion_pago(sender, **kwargs):
    ipn_obj = sender

    # Procesa la notificación de pago y actualiza el estado del carrito
    if ipn_obj.payment_status == 'Completed':
        # Actualiza el estado del carrito a "pagado"
        # Puedes utilizar la información de ipn_obj para identificar y actualizar el carrito correspondiente
        # Por ejemplo: Carrito.objects.filter(id=carrito_id).update(estado_pago='pagado')
        pass

valid_ipn_received.connect(manejar_notificacion_pago)"""